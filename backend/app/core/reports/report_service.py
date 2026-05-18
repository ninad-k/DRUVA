"""Portfolio and trade reports — PDF and Excel.

Provides three report types:
- equity_curve_pdf:        equity curve chart + summary metrics as PDF bytes.
- trade_history_excel:     full trade/order history as a two-sheet Excel workbook.
- portfolio_snapshot_pdf:  current holdings + allocation pie chart as PDF bytes.

PDF is generated with reportlab; charts are produced by matplotlib, serialised
to PNG via BytesIO, and embedded inline.  Excel is generated with openpyxl.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any
from uuid import UUID

import matplotlib
matplotlib.use("Agg")  # non-interactive backend — must be set before importing pyplot
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models.order import Order
from app.db.models.portfolio import PortfolioSnapshot
from app.db.models.position import Position

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
_HEADER_HEX  = "1F4E79"   # dark blue — Word-style corporate header
_ACCENT_HEX  = "2E75B6"   # mid blue
_ALT_HEX     = "D6E4F0"   # light blue row stripe

_RL_HEADER   = colors.HexColor(f"#{_HEADER_HEX}")
_RL_ACCENT   = colors.HexColor(f"#{_ACCENT_HEX}")
_RL_ALT      = colors.HexColor(f"#{_ALT_HEX}")
_RL_WHITE    = colors.white
_RL_BLACK    = colors.black


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _now_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _fig_to_png_bytes(fig: Any) -> bytes:
    """Serialise a matplotlib Figure to PNG bytes and close the figure."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf.read()


def _rl_table_style(row_count: int) -> TableStyle:
    """Standard ReportLab TableStyle with header + alternating row colours."""
    commands = [
        ("BACKGROUND",  (0, 0), (-1, 0),  _RL_HEADER),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  _RL_WHITE),
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0),  9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING",    (0, 0), (-1, 0), 6),
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("GRID",        (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_RL_WHITE, _RL_ALT]),
    ]
    return TableStyle(commands)


def _build_pdf_header(
    story: list,
    styles: Any,
    title: str,
    subtitle: str,
) -> None:
    """Append title + subtitle paragraphs to a ReportLab story list."""
    h1 = ParagraphStyle(
        "H1",
        parent=styles["Heading1"],
        textColor=_RL_HEADER,
        spaceAfter=4,
    )
    h2 = ParagraphStyle(
        "H2",
        parent=styles["Normal"],
        textColor=_RL_ACCENT,
        fontSize=10,
        spaceAfter=12,
    )
    story.append(Paragraph("DHRUVA Portfolio Report", h1))
    story.append(Paragraph(title, h2))
    story.append(Paragraph(subtitle, styles["Normal"]))
    story.append(Spacer(1, 0.5 * cm))


# ---------------------------------------------------------------------------
# Metric calculations
# ---------------------------------------------------------------------------

def _calc_metrics(equity: list[float]) -> dict[str, float]:
    """Compute basic portfolio metrics from an equity curve (list of floats)."""
    if len(equity) < 2:
        return {"total_return": 0.0, "sharpe": 0.0, "max_drawdown": 0.0, "win_rate": 0.0}

    arr    = np.array(equity, dtype=float)
    rets   = np.diff(arr) / arr[:-1]

    total_return = float(arr[-1] / arr[0] - 1) * 100
    sharpe       = float(np.mean(rets) / np.std(rets) * np.sqrt(252)) if np.std(rets) > 0 else 0.0
    running_max  = np.maximum.accumulate(arr)
    drawdowns    = (arr - running_max) / running_max
    max_drawdown = float(drawdowns.min()) * 100
    win_rate     = float(np.sum(rets > 0) / len(rets)) * 100

    return {
        "total_return": round(total_return, 2),
        "sharpe":       round(sharpe, 2),
        "max_drawdown": round(max_drawdown, 2),
        "win_rate":     round(win_rate, 2),
    }


# ---------------------------------------------------------------------------
# 1. Equity curve PDF
# ---------------------------------------------------------------------------

async def equity_curve_pdf(
    account_id: str | UUID,
    period: str,
    session: AsyncSession,
) -> bytes:
    """Generate a PDF containing the equity curve chart and summary metrics.

    Parameters
    ----------
    account_id:
        The account to report on.
    period:
        Human-readable period label (e.g. "2025-01-01 – 2025-12-31").
    session:
        An open AsyncSession.

    Returns
    -------
    Raw PDF bytes suitable for returning as an HTTP response.
    """
    account_uuid = UUID(str(account_id))

    # Fetch portfolio snapshots ordered chronologically.
    rows = (
        await session.execute(
            select(PortfolioSnapshot)
            .where(PortfolioSnapshot.account_id == account_uuid)
            .order_by(PortfolioSnapshot.ts.asc())
        )
    ).scalars().all()

    if rows:
        timestamps = [r.ts for r in rows]
        equity     = [float(r.equity) for r in rows]
    else:
        # Fallback: synthesise a flat curve from order fills (cumulative value).
        orders = (
            await session.execute(
                select(Order)
                .where(
                    Order.account_id == account_uuid,
                    Order.status == "filled",
                )
                .order_by(Order.created_at.asc())
            )
        ).scalars().all()
        if orders:
            timestamps = [o.created_at for o in orders]
            prices     = [float(o.price or o.quantity) for o in orders]
            equity     = [sum(prices[: i + 1]) for i in range(len(prices))]
        else:
            timestamps = [datetime.now(timezone.utc)]
            equity     = [0.0]

    metrics = _calc_metrics(equity)

    # --- Chart -----------------------------------------------------------
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(timestamps, equity, color=f"#{_ACCENT_HEX}", linewidth=1.5)
    ax.fill_between(timestamps, equity, alpha=0.15, color=f"#{_ACCENT_HEX}")
    ax.set_title("Equity Curve", fontsize=12, color=f"#{_HEADER_HEX}")
    ax.set_xlabel("Date", fontsize=9)
    ax.set_ylabel("Portfolio Value (INR)", fontsize=9)
    ax.tick_params(labelsize=8)
    ax.grid(True, linestyle="--", alpha=0.4)
    fig.tight_layout()
    chart_bytes = _fig_to_png_bytes(fig)

    # --- Build PDF -------------------------------------------------------
    buf    = io.BytesIO()
    styles = getSampleStyleSheet()
    doc    = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    story: list = []
    _build_pdf_header(
        story,
        styles,
        title=f"Account: {account_id}",
        subtitle=f"Period: {period}  |  Generated: {_now_str()}",
    )

    # Embed chart
    img_buf = io.BytesIO(chart_bytes)
    img     = Image(img_buf, width=16 * cm, height=6 * cm)
    story.append(img)
    story.append(Spacer(1, 0.5 * cm))

    # Summary table
    summary_data = [
        ["Metric", "Value"],
        ["Total Return", f"{metrics['total_return']:+.2f}%"],
        ["Sharpe Ratio", f"{metrics['sharpe']:.2f}"],
        ["Max Drawdown", f"{metrics['max_drawdown']:.2f}%"],
        ["Win Rate",     f"{metrics['win_rate']:.1f}%"],
    ]
    tbl = Table(summary_data, colWidths=[8 * cm, 8 * cm])
    tbl.setStyle(_rl_table_style(len(summary_data)))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# 2. Trade history Excel
# ---------------------------------------------------------------------------

async def trade_history_excel(
    account_id: str | UUID,
    session: AsyncSession,
) -> bytes:
    """Generate a two-sheet Excel workbook with trade history and per-symbol summary.

    Returns
    -------
    Raw .xlsx bytes.
    """
    account_uuid = UUID(str(account_id))

    orders = (
        await session.execute(
            select(Order)
            .where(Order.account_id == account_uuid)
            .order_by(Order.created_at.asc())
        )
    ).scalars().all()

    wb = Workbook()

    # ---- Sheet 1: Trades ------------------------------------------------
    ws_trades = wb.active
    ws_trades.title = "Trades"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor=_HEADER_HEX)
    alt_fill     = PatternFill("solid", fgColor=_ALT_HEX)
    center_align = Alignment(horizontal="center")

    trade_headers = ["Date", "Symbol", "Side", "Qty", "Price", "Value", "P&L", "Status"]
    ws_trades.append(trade_headers)
    for col_idx, _ in enumerate(trade_headers, start=1):
        cell = ws_trades.cell(row=1, column=col_idx)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = center_align

    for row_num, order in enumerate(orders, start=2):
        qty   = float(order.quantity or 0)
        price = float(order.price or 0)
        value = qty * price
        pnl   = 0.0  # Order-level P&L is not tracked; position carries realized_pnl.

        row_data = [
            order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
            order.symbol,
            str(order.side.value if hasattr(order.side, "value") else order.side),
            qty,
            price,
            round(value, 2),
            round(pnl, 2),
            str(order.status.value if hasattr(order.status, "value") else order.status),
        ]
        ws_trades.append(row_data)

        if row_num % 2 == 0:
            for col_idx in range(1, len(trade_headers) + 1):
                ws_trades.cell(row=row_num, column=col_idx).fill = alt_fill

    _auto_size_columns(ws_trades)

    # ---- Sheet 2: Summary -----------------------------------------------
    ws_summary = wb.create_sheet("Summary")
    summary_headers = ["Symbol", "Total Trades", "Buy Trades", "Sell Trades",
                       "Avg Price", "Total Value", "Realized P&L"]
    ws_summary.append(summary_headers)
    for col_idx, _ in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    # Group by symbol
    grouped: dict[str, list[Order]] = {}
    for order in orders:
        grouped.setdefault(order.symbol, []).append(order)

    # Fetch realized PnL from positions table for accurate per-symbol figure.
    positions = (
        await session.execute(
            select(Position).where(Position.account_id == account_uuid)
        )
    ).scalars().all()
    realized_by_symbol = {p.symbol: float(p.realized_pnl) for p in positions}

    for row_num, (symbol, sym_orders) in enumerate(sorted(grouped.items()), start=2):
        buy_orders  = [o for o in sym_orders if str(getattr(o.side, "value", o.side)).upper() == "BUY"]
        sell_orders = [o for o in sym_orders if str(getattr(o.side, "value", o.side)).upper() == "SELL"]
        prices      = [float(o.price or 0) for o in sym_orders if o.price]
        total_value = sum(float(o.quantity or 0) * float(o.price or 0) for o in sym_orders)
        avg_price   = float(np.mean(prices)) if prices else 0.0
        realized    = realized_by_symbol.get(symbol, 0.0)

        row_data = [
            symbol,
            len(sym_orders),
            len(buy_orders),
            len(sell_orders),
            round(avg_price, 2),
            round(total_value, 2),
            round(realized, 2),
        ]
        ws_summary.append(row_data)

        if row_num % 2 == 0:
            for col_idx in range(1, len(summary_headers) + 1):
                ws_summary.cell(row=row_num, column=col_idx).fill = alt_fill

    _auto_size_columns(ws_summary)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# 3. Portfolio snapshot PDF
# ---------------------------------------------------------------------------

async def portfolio_snapshot_pdf(
    account_id: str | UUID,
    session: AsyncSession,
) -> bytes:
    """Generate a PDF with current holdings and an allocation pie chart.

    Returns
    -------
    Raw PDF bytes.
    """
    account_uuid = UUID(str(account_id))

    positions = (
        await session.execute(
            select(Position).where(Position.account_id == account_uuid)
        )
    ).scalars().all()

    # --- Pie chart -------------------------------------------------------
    labels  = [p.symbol for p in positions] if positions else ["No positions"]
    values  = [float(p.quantity) * float(p.avg_cost) for p in positions] if positions else [1.0]

    fig, ax = plt.subplots(figsize=(6, 6))
    wedge_colors = plt.cm.tab20.colors  # type: ignore[attr-defined]
    ax.pie(
        values,
        labels=labels,
        autopct="%1.1f%%" if sum(values) > 0 else None,
        colors=wedge_colors[: len(labels)],
        startangle=140,
        textprops={"fontsize": 8},
    )
    ax.set_title("Portfolio Allocation", fontsize=12, color=f"#{_HEADER_HEX}")
    chart_bytes = _fig_to_png_bytes(fig)

    # --- Build PDF -------------------------------------------------------
    buf    = io.BytesIO()
    styles = getSampleStyleSheet()
    doc    = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    story: list = []
    _build_pdf_header(
        story,
        styles,
        title=f"Account: {account_id}",
        subtitle=f"Snapshot as of {_now_str()}",
    )

    img_buf = io.BytesIO(chart_bytes)
    img     = Image(img_buf, width=10 * cm, height=10 * cm)
    story.append(img)
    story.append(Spacer(1, 0.5 * cm))

    # Holdings table
    total_value = sum(values) or 1.0
    holdings_data = [["Symbol", "Exchange", "Qty", "Avg Cost", "Value (INR)", "Alloc %", "Realized P&L"]]
    for pos, val in zip(positions, values):
        holdings_data.append([
            pos.symbol,
            str(pos.exchange.value if hasattr(pos.exchange, "value") else pos.exchange),
            str(float(pos.quantity)),
            f"{float(pos.avg_cost):,.2f}",
            f"{val:,.2f}",
            f"{val / total_value * 100:.1f}%",
            f"{float(pos.realized_pnl):,.2f}",
        ])

    if not positions:
        holdings_data.append(["—", "—", "—", "—", "—", "—", "—"])

    col_widths = [3 * cm, 2.5 * cm, 2 * cm, 2.5 * cm, 3 * cm, 2 * cm, 3 * cm]
    tbl = Table(holdings_data, colWidths=col_widths)
    tbl.setStyle(_rl_table_style(len(holdings_data)))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _auto_size_columns(ws) -> None:  # type: ignore[no-untyped-def]
    """Set each column width to fit the widest cell value (max 50 chars)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
            except Exception:  # noqa: BLE001
                cell_len = 0
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 4, 54)


# ---------------------------------------------------------------------------
# Public class wrapper (used by the REST endpoint)
# ---------------------------------------------------------------------------

class ReportService:
    """Thin wrapper so the REST layer can instantiate a single service object."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def equity_curve_pdf(self, account_id: str, period: str) -> bytes:
        return await equity_curve_pdf(account_id, period, self.session)

    async def trade_history_excel(self, account_id: str) -> bytes:
        return await trade_history_excel(account_id, self.session)

    async def portfolio_snapshot_pdf(self, account_id: str) -> bytes:
        return await portfolio_snapshot_pdf(account_id, self.session)
